"""
Billing invoice export — PDF (reportlab) and Excel (openpyxl), both already
project dependencies (see catalogue_render_service.py's PDF use and
requirements.txt). Reads directly off a persisted Sale row (the immutable
snapshot) plus the tenant's own branding fields (name/GST/logo/contact) —
never recomputed, never touches the live gold rate.
"""

import io
from typing import Optional

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from openpyxl import Workbook

from app.models.billing import Sale, Quotation
from app.models.auth import Tenant


def _rows(sale: Sale):
    # Gold Profit is an INTERNAL margin — never itemised on a customer document
    # (invoice PDF / Excel). It is folded into the Gold Value line so the visible
    # rows still reconcile to Subtotal (subtotal_before_tax already includes it).
    return [
        ("Gold Value", sale.gold_value_amount + (sale.gold_profit_amount or 0)),
        (f"Making Charge ({sale.making_charge_type})", sale.making_charge_amount),
        (f"Wastage ({sale.wastage_type})", sale.wastage_amount),
        ("Stone Charge", sale.stone_charge_amount),
        ("Other Charges", sale.other_charges_amount),
        ("Subtotal", sale.subtotal_before_tax),
        (f"GST ({sale.tax_rate_percent}%)" if sale.gst_applied else "GST (not applied)", sale.tax_amount),
        ("Discount", -sale.discount_amount),
        ("Final Amount", sale.final_amount),
    ]


def build_invoice_pdf(sale: Sale, tenant: Optional[Tenant]) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 25 * mm

    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, y, (tenant.name if tenant else "Invoice"))
    c.setFont("Helvetica", 9)
    y -= 6 * mm
    if tenant and tenant.gst_number:
        c.drawString(20 * mm, y, f"GSTIN: {tenant.gst_number}")
        y -= 5 * mm
    if tenant and tenant.contact_phone:
        c.drawString(20 * mm, y, f"Phone: {tenant.contact_phone}")
        y -= 5 * mm

    y -= 4 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, f"Invoice: {sale.invoice_number}")
    c.drawRightString(width - 20 * mm, y, sale.sale_timestamp.strftime("%d-%b-%Y %H:%M"))
    y -= 6 * mm
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, y, f"Customer: {sale.customer_name or sale.customer_id or 'Walk-in'}")
    if sale.customer_phone:
        c.drawRightString(width - 20 * mm, y, sale.customer_phone)
    y -= 8 * mm

    c.setFont("Helvetica-Bold", 10)
    c.drawString(20 * mm, y, f"{sale.product_code} — {sale.product_name}")
    y -= 5 * mm
    c.setFont("Helvetica", 9)
    huid = f" · HUID {sale.huid}" if sale.huid else ""
    vendor = f" · Vendor {sale.vendor_name}" if sale.vendor_name else ""
    c.drawString(
        20 * mm, y,
        f"Purity {sale.purity} · Gross {sale.gross_weight_grams}g · Net {sale.net_gold_weight_grams}g{huid}{vendor}",
    )
    y -= 5 * mm
    c.drawString(20 * mm, y, f"Gold Rate Applied: Rs.{sale.gold_rate_applied:.2f}/g ({sale.gold_rate_source})")
    y -= 10 * mm

    c.line(20 * mm, y, width - 20 * mm, y)
    y -= 8 * mm
    c.setFont("Helvetica", 10)
    for label, value in _rows(sale):
        c.drawString(20 * mm, y, label)
        c.drawRightString(width - 20 * mm, y, f"Rs.{value:,.2f}")
        y -= 6 * mm

    y -= 4 * mm
    c.setFont("Helvetica", 9)
    c.drawString(20 * mm, y, f"Payment: {sale.payment_method} ({sale.payment_status})")
    y -= 15 * mm
    c.setFont("Helvetica-Oblique", 8)
    c.drawString(20 * mm, y, "This is a system-generated invoice.")

    c.showPage()
    c.save()
    return buf.getvalue()


def build_quotation_pdf(quotation: Quotation, tenant: Optional[Tenant]) -> bytes:
    """Customer-facing quotation PDF. Same reportlab path as build_invoice_pdf —
    NO second PDF architecture. Reads the frozen breakdown_json snapshot the
    engine computed at generation time (never recomputed here) plus the scheme
    PREVIEW; writes nothing.

    Customer-facing privacy: Gold Profit is folded into the Gold Value line (so
    the visible rows still reconcile to Subtotal — subtotal_before_tax already
    includes gold profit), and purchase cost / vendor cost / internal margin /
    profit-loss are NEVER drawn. Clearly marked a quotation, not a tax invoice.
    """
    b = quotation.breakdown_json or {}
    scheme_items = (quotation.scheme_breakdown_json or {}).get("items", [])

    def g(key, default=0.0):
        v = b.get(key)
        return v if v is not None else default

    rows = [
        ("Gold Value", g("gold_value_amount") + g("gold_profit_amount")),
        (f"Making Charge ({b.get('making_charge_type', '')})", g("making_charge_amount")),
        (f"Wastage ({b.get('wastage_type', '')})", g("wastage_amount")),
        *([("Stone Charge", g("stone_charge_amount"))] if g("stone_charge_amount") > 0 else []),
        *([("Other Charges", g("other_charges_amount"))] if g("other_charges_amount") > 0 else []),
        ("Subtotal", g("subtotal_before_tax")),
        (f"GST ({g('tax_rate_percent')}%)" if b.get("gst_applied") else "GST (not applied)", g("tax_amount")),
        *([("Discount", -g("discount_amount"))] if g("discount_amount") > 0 else []),
        ("Grand Total", quotation.final_amount),
    ]

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 25 * mm

    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, y, (tenant.name if tenant else "Quotation"))
    c.setFont("Helvetica", 9)
    y -= 6 * mm
    if tenant and tenant.gst_number:
        c.drawString(20 * mm, y, f"GSTIN: {tenant.gst_number}")
        y -= 5 * mm
    if tenant and tenant.contact_phone:
        c.drawString(20 * mm, y, f"Phone: {tenant.contact_phone}")
        y -= 5 * mm

    y -= 4 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, f"Quotation: {quotation.quotation_number}")
    c.drawRightString(width - 20 * mm, y, quotation.created_at.strftime("%d-%b-%Y %H:%M"))
    y -= 6 * mm
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, y, f"Customer: {quotation.customer_name or 'Walk-in'}")
    if quotation.customer_phone:
        c.drawRightString(width - 20 * mm, y, quotation.customer_phone)
    y -= 8 * mm

    c.setFont("Helvetica-Bold", 10)
    c.drawString(20 * mm, y, f"{quotation.product_code} — {quotation.product_name}")
    y -= 5 * mm
    c.setFont("Helvetica", 9)
    net = g("net_gold_weight_grams")
    c.drawString(20 * mm, y, f"Purity {b.get('purity', '')} · Net Gold {net}g")
    y -= 10 * mm

    c.line(20 * mm, y, width - 20 * mm, y)
    y -= 8 * mm
    c.setFont("Helvetica", 10)
    for label, value in rows:
        c.drawString(20 * mm, y, label)
        c.drawRightString(width - 20 * mm, y, f"Rs.{value:,.2f}")
        y -= 6 * mm

    # Scheme preview (read-only — no balance is spent by a quotation).
    if (quotation.scheme_amount_total or 0) > 0:
        for it in scheme_items:
            c.drawString(20 * mm, y, f"Scheme {it.get('enrollment_number', '')}")
            c.drawRightString(width - 20 * mm, y, f"- Rs.{float(it.get('applied_amount', 0)):,.2f}")
            y -= 6 * mm
        c.setFont("Helvetica-Bold", 10)
        c.drawString(20 * mm, y, "Payable After Scheme")
        c.drawRightString(width - 20 * mm, y, f"Rs.{quotation.outstanding_amount:,.2f}")
        y -= 6 * mm
        c.setFont("Helvetica", 10)

    y -= 9 * mm
    c.setFont("Helvetica-Oblique", 8)
    c.drawString(20 * mm, y, "This is a quotation, not a tax invoice. Prices are subject to the prevailing gold rate at the time of sale.")

    c.showPage()
    c.save()
    return buf.getvalue()


def build_invoice_excel(sale: Sale, tenant: Optional[Tenant]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    header = [
        ("Business", tenant.name if tenant else ""),
        ("GSTIN", tenant.gst_number if tenant else ""),
        ("Invoice No", sale.invoice_number),
        ("Date", sale.sale_timestamp.strftime("%d-%b-%Y %H:%M")),
        ("Customer", sale.customer_name or sale.customer_id or "Walk-in"),
        ("Phone", sale.customer_phone or ""),
        ("Product Code", sale.product_code),
        ("Product", sale.product_name),
        ("Vendor", sale.vendor_name or ""),
        ("HUID", sale.huid or ""),
        ("Purity", sale.purity),
        ("Gross Weight (g)", sale.gross_weight_grams),
        ("Net Gold Weight (g)", sale.net_gold_weight_grams),
        ("Gold Rate Applied (Rs/g)", sale.gold_rate_applied),
        ("Gold Rate Source", sale.gold_rate_source),
    ]
    for row in header:
        ws.append(row)

    ws.append([])
    ws.append(["Charge", "Amount (Rs)"])
    for label, value in _rows(sale):
        ws.append([label, value])
    ws.append([])
    ws.append(["Payment Method", sale.payment_method])
    ws.append(["Payment Status", sale.payment_status])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _payment_history_text(payments: list) -> str:
    """Condenses a sale's ledger into one readable cell, e.g.
    "10-Aug-2026 Rs.10000.00 CASH; 14-Aug-2026 Rs.5000.00 UPI". The Admin
    reading the sheet needs the collection story per invoice without a second
    sheet to cross-reference."""
    return "; ".join(
        f"{p.payment_date.strftime('%d-%b-%Y')} Rs.{p.amount:.2f} {p.payment_method}"
        + (f" ({p.reference_no})" if p.reference_no else "")
        for p in payments
    )


def build_sales_history_excel(
    sales: list,
    payments_by_sale: dict,
    tenant: Optional[Tenant],
    period_label: str,
    status_label: str,
    include_internal: bool = True,
) -> bytes:
    """Sales History export — one row per invoice, rendered off the immutable
    Sale snapshots plus the sale_payments ledger. Same openpyxl path as
    build_invoice_excel above (no second export framework), just a list shape
    instead of a single invoice.

    Internal identifiers (sale id, tenant id, inventory item id, customer id,
    created_by) are deliberately omitted — they are meaningless to the Admin
    and needlessly expose internal structure. Purchase cost and profit/loss
    are commercially sensitive: they are included ONLY when include_internal is
    True (Admin/SuperAdmin). For Staff the caller passes include_internal=False
    and those two columns are omitted entirely — mirroring the JSON redaction in
    billing_service._is_privileged so the export cannot leak what the API hides.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales History"

    ws.append([f"{tenant.name if tenant else ''} — Sales History"])
    ws.append(["Period", period_label])
    ws.append(["Payment Status Filter", status_label])
    ws.append(["Invoices", len(sales)])
    ws.append([])

    headers = [
        "Invoice No", "Sale Date", "Customer", "Phone",
        "Product Code", "Product", "Vendor", "Purity",
        "Gross Weight (g)", "Net Gold Weight (g)", "Gold Rate Applied (Rs/g)",
        *(["Purchase Cost (Rs)"] if include_internal else []),
        "Gold Value (Rs)", "Making (Rs)", "Wastage (Rs)",
        "Stone (Rs)", "Other (Rs)", "Subtotal (Rs)", "GST %", "GST (Rs)",
        "Discount (Rs)", "Invoice Total (Rs)", "Amount Paid (Rs)",
        "Outstanding (Rs)", "Payment Status", "Payment History",
        *(["Profit/Loss (Rs)"] if include_internal else []),
    ]
    ws.append(headers)

    total_invoiced = total_collected = total_outstanding = 0.0
    for s in sales:
        paid = float(s.amount_paid or 0)
        outstanding = max(0.0, float(s.final_amount or 0) - paid)
        total_invoiced += float(s.final_amount or 0)
        total_collected += paid
        total_outstanding += outstanding
        ws.append([
            s.invoice_number,
            s.sale_timestamp.strftime("%d-%b-%Y %H:%M"),
            s.customer_name or "Walk-in",
            s.customer_phone or "",
            s.product_code,
            s.product_name,
            s.vendor_name or "",
            s.purity,
            s.gross_weight_grams,
            s.net_gold_weight_grams,
            s.gold_rate_applied,
            *([s.purchase_cost_snapshot if s.purchase_cost_snapshot is not None else ""]
              if include_internal else []),
            s.gold_value_amount,
            s.making_charge_amount,
            s.wastage_amount,
            s.stone_charge_amount,
            s.other_charges_amount,
            s.subtotal_before_tax,
            s.tax_rate_percent if s.gst_applied else 0,
            s.tax_amount,
            s.discount_amount,
            s.final_amount,
            round(paid, 2),
            round(outstanding, 2),
            s.payment_status,
            _payment_history_text(payments_by_sale.get(s.id, [])),
            *([s.estimated_gross_margin if s.estimated_gross_margin is not None else ""]
              if include_internal else []),
        ])

    ws.append([])
    # Totals align under Invoice Total / Amount Paid / Outstanding regardless of
    # whether the two internal columns are present.
    total_col = headers.index("Invoice Total (Rs)")
    totals_row = [""] * len(headers)
    totals_row[0] = "TOTALS"
    totals_row[total_col] = round(total_invoiced, 2)
    totals_row[total_col + 1] = round(total_collected, 2)
    totals_row[total_col + 2] = round(total_outstanding, 2)
    ws.append(totals_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_ca_export_excel(
    sales: list,
    tenant: Optional[Tenant],
    period_label: str,
) -> bytes:
    """CA / accounting export — one row per invoice, restricted to accounting
    fields the project already stores on the immutable Sale snapshot. NO internal
    margin/purchase cost is exposed, and NO GST/HSN/tax field is invented: every
    tax column here (gst_applied, tax_rate_percent, tax_amount) already exists on
    the Sale row exactly as it was priced. Read-only — reads snapshots, writes
    nothing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CA Export"

    ws.append([tenant.name if tenant else "", "", "GSTIN", tenant.gst_number if tenant else ""])
    ws.append(["Accounting Export", period_label])
    ws.append([])

    headers = [
        "Invoice No", "Date", "Customer", "Product Code", "Product", "HUID", "Purity",
        "Net Gold Weight (g)", "Gold Value (Rs)", "Making (Rs)", "Wastage (Rs)",
        "Stone (Rs)", "Other (Rs)", "Subtotal Before Tax (Rs)", "GST Applied",
        "Tax Rate %", "Tax Amount (Rs)", "Discount (Rs)", "Final Amount (Rs)",
        "Sale Status", "Amount Paid (Rs)", "Amount Refunded (Rs)", "Payment Method",
    ]
    ws.append(headers)

    for s in sales:
        ws.append([
            s.invoice_number,
            s.sale_timestamp.strftime("%d-%b-%Y"),
            s.customer_name or s.customer_id or "Walk-in",
            s.product_code,
            s.product_name,
            s.huid or "",
            s.purity,
            s.net_gold_weight_grams,
            s.gold_value_amount,
            s.making_charge_amount,
            s.wastage_amount,
            s.stone_charge_amount,
            s.other_charges_amount,
            s.subtotal_before_tax,
            "Yes" if s.gst_applied else "No",
            s.tax_rate_percent,
            s.tax_amount,
            s.discount_amount,
            s.final_amount,
            s.sale_status,
            s.amount_paid,
            s.amount_refunded,
            s.payment_method,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_ca_export_excel(
    sales: list,
    tenant: Optional[Tenant],
    period_label: str,
) -> bytes:
    """CA / accounting export — one row per invoice, restricted to accounting
    fields the project already stores on the immutable Sale snapshot. NO internal
    margin/purchase cost is exposed, and NO GST/HSN/tax field is invented: every
    tax column here (gst_applied, tax_rate_percent, tax_amount) already exists on
    the Sale row exactly as it was priced. Read-only — reads snapshots, writes
    nothing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CA Export"

    ws.append([tenant.name if tenant else "", "", "GSTIN", tenant.gst_number if tenant else ""])
    ws.append(["Accounting Export", period_label])
    ws.append([])

    headers = [
        "Invoice No", "Date", "Customer", "Product Code", "Product", "HUID", "Purity",
        "Net Gold Weight (g)", "Gold Value (Rs)", "Making (Rs)", "Wastage (Rs)",
        "Stone (Rs)", "Other (Rs)", "Subtotal Before Tax (Rs)", "GST Applied",
        "Tax Rate %", "Tax Amount (Rs)", "Discount (Rs)", "Final Amount (Rs)",
        "Sale Status", "Amount Paid (Rs)", "Amount Refunded (Rs)", "Payment Method",
    ]
    ws.append(headers)

    for s in sales:
        ws.append([
            s.invoice_number,
            s.sale_timestamp.strftime("%d-%b-%Y"),
            s.customer_name or s.customer_id or "Walk-in",
            s.product_code,
            s.product_name,
            s.huid or "",
            s.purity,
            s.net_gold_weight_grams,
            s.gold_value_amount,
            s.making_charge_amount,
            s.wastage_amount,
            s.stone_charge_amount,
            s.other_charges_amount,
            s.subtotal_before_tax,
            "Yes" if s.gst_applied else "No",
            s.tax_rate_percent,
            s.tax_amount,
            s.discount_amount,
            s.final_amount,
            s.sale_status,
            s.amount_paid,
            s.amount_refunded,
            s.payment_method,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
